
from django.contrib import auth
from .models import *
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from openpyxl import Workbook
from django.http import HttpResponse 

# Create your views here.


@login_required(login_url='login')
def index(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        if user is not None:
            request.session['userlogin'] = True
            auth.login(request, user)
            return redirect('index')
        else:
            messages.info(request, 'Check Username and Password')
            return redirect('login')
    return render(request, "index.html")


def signup(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        if password1 == password2:
            if User.objects.filter(username=username).exists():
                messages.info(request, "Username Exists")
                return redirect('signup')
            elif User.objects.filter(email=email).exists():
                messages.info(request, "Email id Exits")
                return redirect('signup')
            else:
                user = User.objects.create_user(
                    username=username, password=password1, email=email)
                user.save()
    return render(request, "signup.html")


# def signupp(request):
#     if request.method == "POST":
#         full_name = request.POST.get("full_name")
#         email = request.POST.get("email")
#         password = request.POST.get("password")
#         confirm_password = request.POST.get("confirm_password")
#         # phone = request.POST.get("phone")
#         # dob = request.POST.get("dob")
#         # gender = request.POST.get("gender")
#         # message_text = request.POST.get("message")
#         # checkbox returns 'on' if checked
#         agree = request.POST.get("agree") == "on"

#         # ✅ Check if email already exists
#         if Person.objects.filter(email=email).exists():
#             messages.error(request, "❌ This email is already registered!")
#             return redirect("signup")

#         # ✅ Save to database
#         Person.objects.create(
#             full_name=full_name,
#             email=email,
#             password=password,
#             confirm_password=confirm_password,
#             # phone=phone,
#             # dob=dob,
#             # gender=gender,
#             # message=message_text,
#             agree_terms=agree
#         )

#         messages.success(request, "✅ Form submitted successfully!")
#         # redirect to homepage (or person-list if you want)
#         return redirect("signup")

#     return render(request, "signup.html")


# def loginn(request):
#     if request.method == "POST":
#         username = request.POST.get("username")
#         password = request.POST.get("password")

#         user = authenticate(request, username=username, password=password)

#         if user is not None:
#             login(request, user)
#             messages.success(request, "Login successful ✅")
#             return redirect("index")  # redirect to your home page
#         else:
#             messages.error(request, "Invalid username or password ❌")

#     return render(request, "login.html")


def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        if user is not None:
            request.session['userlogin'] = True
            auth.login(request, user)
            return redirect('index')
        else:
            messages.info(request, 'Check Username and Password')
            return redirect('login')
    return render(request, 'login.html')


def forms(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        dob = request.POST.get("dob")
        gender = request.POST.get("gender")
        message = request.POST.get("message")
        agree = request.POST.get("agree") == "on"

        # Save to database
        person = Forms.objects.create(
            full_name=full_name,
            email=email,
            phone=phone,
            dob=dob,
            gender=gender,
            message=message,
            agree_terms=agree
        )
        person.save()

        messages.success(request, "Form submitted successfully ✅")
        return redirect("forms")  # redirect to a list page after submission

    return render(request, 'forms.html')


def table(request):
    list = Forms.objects.all()

    return render(request, "table.html", {"list": list})


def table_edit(request, id):
    edit = Forms.objects.filter(id=id)
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        dob = request.POST.get("dob")
        gender = request.POST.get("gender")
        message = request.POST.get("message")
        # agree = request.POST.get("agree") == "on"
        editform = Forms.objects.filter(id=id).update(full_name=full_name,email=email,phone=phone,dob=dob,gender=gender,message=message,)

    return render(request, "table_edit.html", {"edit": edit})

from django.shortcuts import get_object_or_404, redirect

def table_delete(request, id):
    # form = get_object_or_404(Forms, id=id) 
    form = Forms.objects.filter(id=id)  # find record or 404
    form.delete()                            # delete from DB
    return redirect("table")                 # redirect to your table page


def logout(request):
    request.session.clear()
    return redirect('index')



@login_required(login_url='login')
def downloadtotalrecord_excel(request):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Sn'
    ws['B1'] = 'Full Name'
    ws['C1'] = 'Email'
    ws['D1'] = 'Phone'
    ws['E1'] = 'Date of Birth'
    ws['F1'] = 'Gender'
    ws['G1'] = 'Message'
    ws['H1'] = 'Agree'
    
    row_number = 2
    index=1
    employees=Forms.objects.all()
    for i in employees:
        ws.cell(row=row_number, column=1, value=index)
        ws.cell(row=row_number, column=2, value=i.full_name)
        ws.cell(row=row_number, column=3, value=i.email)  # Add other fields as needed
        ws.cell(row=row_number, column=4, value=i.phone)
        ws.cell(row=row_number, column=5, value=i.dob)
        ws.cell(row=row_number, column=6, value=i.gender)
        ws.cell(row=row_number, column=7, value=i.message)
        ws.cell(row=row_number, column=8, value=i.agree_terms)
        
        row_number += 1
        index=index+1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_excel_file.xlsx'
    wb.save(response)
    return response